# Code to take an excel list and assign the corresponding template
# Uses a thread pool for parallel Agile SOAP calls, with a per-thread
# zeep client (SOAP clients are not thread-safe) and a shared description
# cache to avoid redundant calls.

import os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from zeep import Client, Settings, helpers
from zeep.transports import Transport
from requests import Session
from requests.auth import HTTPBasicAuth
from lxml import etree
from pathlib import Path
import sqlite3

# ─────────────────────────── CONFIG ────────────────────────────────────────

ROOT_FOLDER   = os.path.abspath(os.path.dirname(__file__))
import_excel  = os.path.join(ROOT_FOLDER, "Import_WD.xlsx")

AGILE_TABLE_WSDL = "http://pagapps1.ichorsystems.com:7001/CoreService/services/Table?wsdl"
DB_PATH = Path(r"\\sg1sfile5\dept$\Engineering\029 ECO SG\WD\automation.db")

USERNAME = os.getenv("AGILE_USER")
PASSWORD = os.getenv("AGILE_PASS")

MAX_WORKERS = 8  # tune to how many parallel connections Agile tolerates

#DB Mapping
SHEET_CONFIG = {

    "Work Definition Headers": {
        "table": "wd_headers_production",
        "db_template_column": "template_name",
        "column_mapping": {
            "Name": "template_name",
            "Action Code": "action_code",
            "Organization Code": "org_code",
            "Work Definition Status": "wd_status",
            "Structure Internal Name": "wd_structure_name",
            "Work Definition Internal Name": "wd_internal_name",
            "Version": "version",
            "Production Priority": "production_priority",
            "Serial Tracked": "serial_tracked",
            "Completion Subinventory": "completion_subinv",
            " Completion Locator Segment1": "completion_locator"
        }
    },

    "Work Definition Operations": {
        "table": "wd_operations_production",
        "db_template_column": "template_name",
        "column_mapping": {
            "Name": "template_name",
            "Operation Sequence": "operation_sequence",
            "Standard Operation Code": "standard_operation_code",
            "Action Code": "action_code",
            "Referenced": "is_referenced",
            "Operation Type": "operation_type",
            "Count Point ": "count_point",
            "Automatically Transact": "automatically_transact",
            "Serialization": "serialization"
        }
    },

    "Operation Items - Standard": {
        "table": "wd_items",
        "db_template_column": "template_name",
        "column_mapping": {
            "Name": "template_name",
            "Action Code": "action_code",
            "Find Number": "find_number",
            "Basis": "basis",
            "Ad hoc item": "ad_hoc",
            "Supply Subinventory": "supply_subinv",
            "Supply Locator Segment1": "supply_loc_seg1"
        }
    },

    "Operation Resources": {
        "table": "wd_resources_production",
        "db_template_column": "name",
        "column_mapping": {
            "Name": "name",
            "Action Code": "action_code",
            "Operation Sequence": "operation_sequence",
            "Resource Sequence": "resource_sequence",
            "Resource Code": "operation_code_key",
            "Units Assigned": "units_assigned",
            "Basis": "basis",
            "Usage": "usage",
            "Scheduled": "scheduled",
            "Principal": "principal",
            "Charge Type": "charge_type"
        }
    }
}

# ─────────────────────────── TEMPLATE MAP ──────────────────────────────────

template_dict = {
    'A77267': 'SGP_TEMPLATE_LAM_EOS_GB',
    '270980': 'SGP_TEMPLATE_LAM_EOS_DA_DSL',
    '317601': 'SGP_TEMPLATE_LAM_EOS_DA_ESX',
    '290615': 'SGP_TEMPLATE_LAM_EOS_DA_ESX',
    '318645': 'SGP_TEMPLATE_LAM_EOS_DA_KIT',

    '247675': 'SGP_TEMPLATE_LAM_EOS_INTCON',
    '247676': 'SGP_TEMPLATE_LAM_EOS_INTCON',

    '3344':   'SGP_TEMPLATE_LAM_EOS_DA_MOD',

    '113811': 'SGP_TEMPLATE_LAM_UFA_KIT',

    '068755': 'SGP_TEMPLATE_LAM_GB_JTS_KIT',
    '277633': 'SGP_TEMPLATE_LAM_GB_JTS2',
    '246154': 'SGP_TEMPLATE_LAM_GB_JTS2_KIT',

    '152659': 'SGP_TEMPLATE_LAM_ICS',
    '257153': 'SGP_TEMPLATE_LAM_ZONE',

    'A47685': 'SGP_TEMPLATE_LAM_SENSEI_FIP',
    'A32849': 'SGP_TEMPLATE_LAM_SENSEI_BRIDGE',
    '314500': 'SGP_TEMPLATE_LAM_SENSEI_FIB',

    'A49959': 'SGP_TEMPLATE_LAM_SENSEI_PF_L6',
    'A32811': 'SGP_TEMPLATE_LAM_SENSEI_PF_L6',

    'A36987': 'SGP_TEMPLATE_LAM_ARGOS',
    'A36387': 'SGP_TEMPLATE_LAM_MOBI',

    '290639': 'SGP_TEMPLATE_LAM_NPI',
    '051190': 'SGP_TEMPLATE_LAM_AFVI_GB',
    '277635': 'SGP_TEMPLATE_LAM_AFVI2_JTS',

    'A22849': 'SGP_TEMPLATE_LAM_WITH_853-051190',
    'A47482': 'SGP_TEMPLATE_LAM_VDS',

    '067200': 'SGP_TEMPLATE_LAM_ARGOS',
    '079849': 'SGP_LAM_AFVI_WITH_853-051190',
    
    '047693': 'SGP_LAM_AFVI_047693',
    '313217': 'SGP_LAM_AFVI_313217',
    
    '290629': 'SGP_TEMPLATE_LAM_EOS_DA_MDL',
    '318645': 'SGP_TEMPLATE_LAM_EOS_DA_MDL',
    'B03058': 'SGP_TEMPLATE_LAM_EOS_DA_MDL'

}

# ─────────────────────────── PER-THREAD CLIENT ─────────────────────────────
# zeep Client objects are not thread-safe, so each thread gets its own
# via threading.local().

_thread_local = threading.local()

def get_client():
    """Return a zeep Client and its bound types for the current thread."""
    if not hasattr(_thread_local, "client"):
        session = Session()
        session.auth = HTTPBasicAuth(USERNAME, PASSWORD)
        transport = Transport(session=session)
        settings  = Settings(strict=False, xml_huge_tree=True)
        c = Client(AGILE_TABLE_WSDL, transport=transport, settings=settings)
        _thread_local.client              = c
        _thread_local.RequestTableType   = c.get_type(
            "{http://xmlns.oracle.com/AgileObjects/Core/Table/V1}RequestTableType"
        )
        _thread_local.LoadTableRequestType = c.get_type(
            "{http://xmlns.oracle.com/AgileObjects/Core/Table/V1}LoadTableRequestType"
        )
    return (
        _thread_local.client,
        _thread_local.RequestTableType,
        _thread_local.LoadTableRequestType,
    )

# ─────────────────────────── DESCRIPTION CACHE ─────────────────────────────
# Many WDs share child parts, so cache descriptions to avoid duplicate calls.

_desc_cache      = {}
_desc_cache_lock = threading.Lock()

# ─────────────────────────── SOAP HELPERS ──────────────────────────────────

def extract_cell_value(row: dict, names: tuple):
    for elem in row.get("_value_1", []):
        tag = etree.QName(elem.tag).localname.lower()
        if tag in names:
            return elem.text
    return None


def load_bom(parent_part):
    client, RequestTableType, LoadTableRequestType = get_client()
    req = RequestTableType(
        classIdentifier="Part",
        objectNumber=parent_part,
        tableIdentifier="BOM"
    )
    try:
        data = helpers.serialize_object(
            client.service.loadTable(
                LoadTableRequestType(tableRequest=[req])
            )
        )
    except Exception as e:
        print(f"[ERROR] Failed to load BOM for {parent_part}: {e}")
        return []

    rows = data["tableContents"][0].get("row", [])
    results = []
    for r in rows:
        child_part = r["objectReferentId"]["objectName"]
        qty = extract_cell_value(r, ("quantity", "qty", "bomquantity", "quantityper"))
        results.append({"child": child_part, "quantity": qty})
    return results


def get_description(part):
    # Check cache first (no network call needed if already fetched)
    with _desc_cache_lock:
        if part in _desc_cache:
            return _desc_cache[part]

    client, RequestTableType, LoadTableRequestType = get_client()
    req = RequestTableType(
        classIdentifier="Part",
        objectNumber=part,
        tableIdentifier="Title Block"
    )
    try:
        data = helpers.serialize_object(
            client.service.loadTable(
                LoadTableRequestType(tableRequest=[req])
            )
        )
    except Exception:
        with _desc_cache_lock:
            _desc_cache[part] = None
        return None

    desc = None
    rows = data["tableContents"][0].get("row", [])
    for r in rows:
        for elem in r.get("_value_1", []):
            tag = etree.QName(elem.tag).localname.lower()
            if tag in ("description", "desc"):
                desc = elem.text
                break
        if desc:
            break

    with _desc_cache_lock:
        _desc_cache[part] = desc
    return desc

# ─────────────────────────── LOGIC HELPERS ─────────────────────────────────

def count_mfc_sccm(wd):
    total_qty = 0
    level1 = load_bom(wd)

    for item in level1:
        desc = get_description(item["child"])
        if desc and "MFC" in desc and "SCCM" in desc:
            try:
                total_qty += float(item["quantity"])
            except Exception:
                pass

    if total_qty == 0:
        print(f"{wd}: No MFC found at level 1, checking level 2")
        for item in level1:
            for sub in load_bom(item["child"]):
                desc = get_description(sub["child"])
                if desc and "MFC" in desc and "SCCM" in desc:
                    try:
                        total_qty += float(sub["quantity"])
                    except Exception:
                        pass

    return total_qty


def detect_afvi2_type(wd):
    for item in load_bom(wd):
        desc = get_description(item["child"])
        if desc and "CONFIG ASSY" in desc and "AFVI-2" in desc:
            desc_upper = desc.upper()
            if "SINGLE" in desc_upper:
                return "SINGLE"
            elif "DUAL" in desc_upper:
                return "DUAL"
    return None

# ─────────────────────────── PER-ROW WORKER ────────────────────────────────

def process_row(row_num, wd):
    """
    Runs in a worker thread. Returns (row_num, template_value, log_message).
    All Agile calls happen here; Excel writes happen only in the main thread.
    """
    try:
        parts = wd.split('-')
        if len(parts) < 2:
            return row_num, None, f"{wd}: unexpected format, skipping."
        keyword = parts[1]

        if keyword in ("033051", "065780"):
            mfc_qty = count_mfc_sccm(wd)
            if mfc_qty > 16:
                template = 'SGP_TEMPLATE_LAM_GB_JTS_68'
            elif mfc_qty == 16:
                template = 'SGP_TEMPLATE_LAM_GB_JTS_60'
            elif mfc_qty >= 12:
                template = 'SGP_TEMPLATE_LAM_GB_JTS_52'
            elif mfc_qty == 0:
                template = None
            else:
                template = 'SGP_TEMPLATE_LAM_GB_JTS_48'
            log = f"{wd} → {mfc_qty} MFC SCCM parts"
            if mfc_qty == 0:
                log = f"{wd}: Found no sticks inside"
            return row_num, template, log

        elif keyword == "313218":
            afvi_type = detect_afvi2_type(wd)
            if afvi_type == "SINGLE":
                template = "SGP_TEMPLATE_LAM_AFVI2_GB_SINGLE"
            elif afvi_type == "DUAL":
                template = "SGP_TEMPLATE_LAM_AFVI2_GB_DUAL"
            else:
                template = None
            log = (
                f"{wd} → AFVI2 type: {afvi_type}"
                if afvi_type
                else f"{wd}: No AFVI-2 CONFIG ASSY found or type unclear"
            )
            return row_num, template, log

        elif keyword in ("035344", "082521", "277632"):
            desc = get_description(wd)
            template = None
            if desc:
                if "E4" in desc:
                    template = "SGP_TEMPLATE_LAM_UFA_E4"
                elif "E5" in desc:
                    template = "SGP_TEMPLATE_LAM_UFA_E5"
                elif "E6" in desc:
                    template = "SGP_TEMPLATE_LAM_UFA_E6"
            log = (
                f"{wd} description: {desc}"
                if template
                else f"{wd}: description does not contain E4/E5/E6"
            )
            return row_num, template, log

        elif keyword in template_dict:
            template = template_dict[keyword]
            return row_num, template, f"{template} assigned to {wd}"

        else:
            return row_num, None, f"{wd}: ignored due to non-matching template."

    except Exception as e:
        return row_num, None, f"[ERROR] {wd} row {row_num}: {e}"

# ─────────────────────────── MAIN ──────────────────────────────────────────

def main_template():
    wb = load_workbook(import_excel)
    ws = wb['Sheet1']

    # Collect all rows that still need a template (column B is empty)
    pending = []
    for r in range(2, ws.max_row + 1):
        if ws[f'B{r}'].value is not None:
            continue
        wd = ws[f'A{r}'].value
        if wd:
            pending.append((r, str(wd).strip()))

    print(f"Processing {len(pending)} rows with {MAX_WORKERS} workers...")

    results = {}  # row_num → template value

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(process_row, row_num, wd): (row_num, wd)
            for row_num, wd in pending
        }
        for future in as_completed(futures):
            row_num, template, log = future.result()
            print(log)
            results[row_num] = template

    # Write results back to Excel on the main thread (openpyxl is not thread-safe)
    for row_num, template in results.items():
        if template is not None:
            ws[f'B{row_num}'].value = template

    wb.save(import_excel)
    print(f"Done. {sum(1 for v in results.values() if v)} templates assigned.")

def main_quote():
    main_template()
    wb = load_workbook(import_excel)
    ws = wb['Sheet1']
    pending = {}
    #Collect all empty rows
    for r in range(2, ws.max_row + 1):
        template = ws[f'B{r}'].value
        if template:
            #Connects to the WD db to collect the hours
            conn = sqlite3.connect(DB_PATH)
            #Finds all the rows of the template without the leadtime, and sums them
            try:
                cursor = conn.execute(
                    """
                    SELECT SUM(usage)
                    FROM wd_resources_production
                    WHERE name LIKE ?
                    AND operation_code_key != 'LEADTIME_DAYS'
                    """,
                    (template,)
                )
                result = cursor.fetchone()[0]
                print(f"{result} hours for {template}")
                ws[f'B{r}'].value = str(result)
            except Exception as e:
                print(f"Error has occured: {e}")
    wb.save(import_excel)
    
    
if __name__ == "__main__":
    choice = None
    while choice not in (1, 2):
        try:
            choice = int(input("Set 1 for template, set 2 for quotation: "))
            if choice not in (1, 2):
                print("Invalid option. Try again.")
        except ValueError:
            print("Error, not a number.")
    if choice == 1:
        main_template()
    elif choice == 2:
        #Does the main template then converts them into hours
        main_quote()
