import os
import requests
from pathlib import Path
from openpyxl import load_workbook
from zeep import Client, Settings, helpers
from zeep.transports import Transport
from requests import Session
from requests.auth import HTTPBasicAuth
from lxml import etree

#----------------------------------------------------
# Configuration
#----------------------------------------------------
ROOT_FOLDER = Path(os.path.abspath(os.path.dirname(__file__)))
JSON_HEADERS = {'Content-Type': 'application/json'}
BOM_INFO_URL = 'http://172.20.52.26/agile-graph-dev/get-bom-info'
input_file = ROOT_FOLDER / "Import_WD.xlsx"
template_file = ROOT_FOLDER / "Core" / "Template.xlsm"
import_template_file = ROOT_FOLDER / "Core" / "WorkDefinitionTemplate_26B.xlsm"
AGILE_WSDL = "http://pagapps1.ichorsystems.com:7001/CoreService/services/Table?wsdl"
AGILE_USER = "kmageshkumar"
AGILE_PASS = "P@ssw0rd2026c"


# Static SOAP client (load once)
_session = Session()
_session.auth = HTTPBasicAuth("kmageshkumar", "P@ssw0rd2026c")
_transport = Transport(session=_session)
_settings = Settings(strict=False, xml_huge_tree=True)

_agile_client = Client(AGILE_WSDL, transport=_transport, settings=_settings)

RequestTableType = _agile_client.get_type(
    "{http://xmlns.oracle.com/AgileObjects/Core/Table/V1}RequestTableType"
)
LoadTableRequestType = _agile_client.get_type(
    "{http://xmlns.oracle.com/AgileObjects/Core/Table/V1}LoadTableRequestType"
)


#---------------HELPER FUNCTIONS----------------
def write_batch_to_import_file(ws_import_batches, batch_code: str):
    ws_import_batches["B6"] = batch_code
    ws_import_batches["C6"] = batch_code

def write_work_definition_headers(ws_input, ws_template_headers, ws_output_headers, batch_code: str):

    # -----------------------
    # Read WD list from input
    # -----------------------
    wd_list = []
    row = 2
    while True:
        wd_num = ws_input[f"A{row}"].value
        tmpl = ws_input[f"B{row}"].value
        if wd_num is None:
            break
        wd_list.append((str(wd_num).strip(), str(tmpl).strip()))
        row += 1

    if not wd_list:
        raise ValueError("No WDs found in Import_WD.xlsx")

    max_cols = ws_template_headers.max_column
    write_row = 6  # output insertion row

    # -----------------------
    # Process each WD
    # -----------------------
    for wd_num, tmpl in wd_list:

        for r in ws_template_headers.iter_rows(min_row=2, max_col=max_cols, values_only=True):
            template_name = r[0]  # column A

            if template_name != tmpl:
                continue

            row_data = list(r[1:])  # columns B → end

            # Replace placeholders
            for i, cell in enumerate(row_data):
                if isinstance(cell, str):
                    cell = cell.replace("__BATCH__", batch_code)
                    cell = cell.replace("__WD__", wd_num)

                    row_data[i] = cell

            # Write into output sheet
            for col_index, value in enumerate(row_data, start=2):
                ws_output_headers.cell(row=write_row, column=col_index, value=value)

            write_row += 1

def write_work_definition_operations(
    ws_input,
    ws_template_ops,
    ws_output_ops,
    batch_code: str,
    start_row: int
):
    """
    Fills the Work Definition Operations sheet.

    - Reads WDs from input file
    - For each WD:
        • finds rows in the template where column A = template_name
        • copies columns B → end
        • replaces placeholders BATCH and WD
        • appends them to the output sheet starting at start_row
    """

    # ------ Read WD list from input ------
    wd_list = []
    row = 2
    while True:
        wd_num = ws_input[f"A{row}"].value
        tmpl = ws_input[f"B{row}"].value
        if wd_num is None:
            break
        wd_list.append((str(wd_num).strip(), str(tmpl).strip()))
        row += 1

    if not wd_list:
        raise ValueError("No WDs found in Import_WD.xlsx")

    max_cols = ws_template_ops.max_column
    write_row = start_row

    # ------ Process each WD ------
    for wd_num, tmpl in wd_list:

        for r in ws_template_ops.iter_rows(min_row=2, max_col=max_cols, values_only=True):
            template_name = r[0]  # Column A marker

            if template_name != tmpl:
                continue

            # Extract columns B onward
            row_data = list(r[1:])

            # Replace placeholders
            for i, cell in enumerate(row_data):
                if isinstance(cell, str):
                    cell = cell.replace("__BATCH__", batch_code)
                    cell = cell.replace("__WD__", wd_num)
                    row_data[i] = cell

            # Write the row to the output sheet
            for col_index, value in enumerate(row_data, start=2):
                ws_output_ops.cell(row=write_row, column=col_index, value=value)

            write_row += 1

def write_operation_resources(
    ws_input,
    ws_template_res,
    ws_output_res,
    batch_code: str,
    start_row: int
):
    """
    Fills the Operation Resources sheet.

    - Reads WDs from input file
    - For each WD:
        • finds rows in the template where column A = template_name
        • copies columns B onward exactly
        • replaces placeholders BATCH and WD
        • writes into the output sheet starting at start_row

    """

    # ------ Read WD entries ------
    wd_list = []
    row = 2
    while True:
        wd_num = ws_input[f"A{row}"].value
        tmpl = ws_input[f"B{row}"].value
        if wd_num is None:
            break
        wd_list.append((str(wd_num).strip(), str(tmpl).strip()))
        row += 1

    if not wd_list:
        raise ValueError("No WDs found in Import_WD.xlsx")

    max_cols = ws_template_res.max_column
    write_row = start_row

    # ------ Process template rows for each WD ------
    for wd_num, tmpl in wd_list:

        for r in ws_template_res.iter_rows(min_row=2, max_col=max_cols, values_only=True):
            template_name = r[0]  # Column A identifies the template

            if template_name != tmpl:
                continue

            # Extract columns B → end
            row_data = list(r[1:])

            # Replace placeholders
            for i, cell in enumerate(row_data):
                if isinstance(cell, str):
                    cell = cell.replace("__BATCH__", batch_code)
                    cell = cell.replace("__WD__", wd_num)
                    row_data[i] = cell

            # Write row to output sheet
            for col_index, value in enumerate(row_data, start=2):
                ws_output_res.cell(row=write_row, column=col_index, value=value)

            write_row += 1

def _extract_text(elem):
    """Return text content from XML element."""
    if elem is None:
        return ""
    return "".join(elem.itertext()).strip()


def get_bom_children(parent_item_number: str):
    """
    Get Level-1 BOM children directly from Agile WSDL.

    Returns list of tuples:
         [ (child_part_number, qty), ... ]
    """

    # Build request
    table_request = RequestTableType(
        classIdentifier="Part",
        objectNumber=parent_item_number,
        tableIdentifier="BOM"
    )

    load_request = LoadTableRequestType(tableRequest=[table_request])

    try:
        response = _agile_client.service.loadTable(load_request)
        data = helpers.serialize_object(response)

    except Exception as e:
        print(f"[ERROR] SOAP BOM load failed for {parent_item_number}: {e}")
        return []

    try:
        rows = data["tableContents"][0]["row"]
    except Exception:
        print(f"[WARN] No BOM table returned for {parent_item_number}")
        return []

    results = []

    # Parse each row
    for row in rows:

        info = row.get("additionalRowInfo", {})
        level = info.get("level", None)

        # Only level 1 components
        if level != 1:
            continue

        # Child item number
        child = row["objectReferentId"]["objectName"]

        qty = ""
        for elem in row.get("_value_1", []):
            tag = etree.QName(elem.tag).localname
            if tag == "qty":
                qty = _extract_text(elem)

        results.append((child, qty))

    return results

def write_operation_items_standard(
    ws_input,
    ws_template_items,
    ws_output_items,
    batch_code: str,
    start_row: int
):
    """
    Fills the 'Operation Items - Standard' sheet.

    For each WD in the input:
      - Fetch BOM via get_bom_children()
      - For each child, copy the template row (columns B → end)
      - Replace:
          __BATCH__
          __WD__
          __ITEM__   (child pn)
          __QTY__    (child qty)
          __SEQUENCE__   (10, 20, 30...)
    """

    # ------------------------
    # Read WD list
    # ------------------------
    wd_list = []
    row = 2
    while True:
        wd_num = ws_input[f"A{row}"].value
        tmpl = ws_input[f"B{row}"].value
        if wd_num is None:
            break
        wd_list.append((str(wd_num).strip(), str(tmpl).strip()))
        row += 1

    if not wd_list:
        raise ValueError("No WDs found in Import_WD.xlsx")

    # Template row count
    max_cols = ws_template_items.max_column

    write_row = start_row

    # ------------------------
    # Process each WD
    # ------------------------
    for wd_num, tmpl in wd_list:

        # Find **the single template row** for this template
        template_row = None

        for r in ws_template_items.iter_rows(min_row=2, max_col=max_cols, values_only=True):
            template_name = r[0]   # column A

            if template_name == tmpl:
                template_row = r
                break

        if template_row is None:
            print(f"[WARN] No Operation Items template for '{tmpl}' — skipping WD {wd_num}")
            continue

        base_row_data = list(template_row[1:])  # columns B → end

        # ------------------------
        # Fetch BOM children for this WD
        # ------------------------
        children = get_bom_children(wd_num)

        if not children:
            print(f"[WARN] No BOM returned for {wd_num}")
            continue

        # ------------------------
        # Insert 10, 20, 30 sequences
        # ------------------------
        sequence = 10

        for (child_pn, child_qty) in children:

            row_data = []
            for cell in base_row_data:
                if isinstance(cell, str):

                    cell = cell.replace("__BATCH__", batch_code)
                    cell = cell.replace("__WD__", wd_num)
                    cell = cell.replace("__ITEM__", child_pn)
                    cell = cell.replace("__QTY__", str(child_qty))
                    cell = cell.replace("__SEQUENCE__", str(sequence))

                row_data.append(cell)

            # Write row to output sheet
            for col_index, value in enumerate(row_data, start=2):
                ws_output_items.cell(row=write_row, column=col_index, value=value)

            write_row += 1
            sequence += 10


#----------------------------------------------------
# MAIN
#----------------------------------------------------
if __name__ == "__main__":
    batch_code = input("Input your unique batch code. This will be used to identify the file in oracle: ").strip()

    # -----------------------------
    # Load ALL workbooks once
    # -----------------------------
    wb_input = load_workbook(input_file) #User input
  
    wb_template = load_workbook(template_file) #Template to draw from

    wb_import = load_workbook(import_template_file, keep_vba=True) #The import file, keep vba is true to preserve the macro

    ws_input = wb_input.active

    # Sheets
    ws_import_batches = wb_import["Import Batches"]
    ws_template_headers = wb_template["Work Definition Headers"]
    ws_output_headers = wb_import["Work Definition Headers"]

    # -----------------------------
    # Execute sheet-specific modules
    # -----------------------------
    write_batch_to_import_file(ws_import_batches, batch_code)

    write_work_definition_headers(
        ws_input,
        ws_template_headers,
        ws_output_headers,
        batch_code
    )

    ws_template_ops = wb_template["Work Definition Operations"]
    ws_output_ops = wb_import["Work Definition Operations"]

    write_work_definition_operations(
        ws_input,
        ws_template_ops,
        ws_output_ops,
        batch_code,
        start_row=6
    )

    ws_template_res = wb_template["Operation Resources"]
    ws_output_res = wb_import["Operation Resources"]

    write_operation_resources(
        ws_input=ws_input,
        ws_template_res=ws_template_res,
        ws_output_res=ws_output_res,
        batch_code=batch_code,
        start_row=6
    )

    ws_template_items = wb_template["Operation Items - Standard"]
    ws_output_items = wb_import["Operation Items - Standard"]

    write_operation_items_standard(
        ws_input=ws_input,
        ws_template_items=ws_template_items,
        ws_output_items=ws_output_items,
        batch_code=batch_code,
        start_row=6
    )


    # -----------------------------
    # SAVE AS new batch file
    # -----------------------------
    output_path = ROOT_FOLDER / f"{batch_code}.xlsm"
    wb_import.save(output_path)

    # -----------------------------
    # CLOSE original template WITHOUT saving
    # -----------------------------
    wb_import.close()

    print(f"\n✔ Successfully created: {output_path}")
